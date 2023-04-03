# -*- coding: utf-8 -*-
"""
Created on Fri Mar 31 14:51:35 2023

@author: jdi004
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import load_workbook
plt.rcParams["font.family"] = "Times New Roman"

filename        = r'xlsx_files/twoloops_results_MATLAB_HDQ_individual.xlsx'

feature         = [    'd1'
#    'p1','p2','p3','p4','p5','p6',
#    'd1','d2','d3','d4','d5','d6',
#    'q1','q2','q3','q4','q5','q6','q7','q8',
    ]
feature_type    = 'Pressure head'
units           = '(wmc)'

models = [
            3

#        0,1,2,3,4
    ]

sp      = [
       24,48,72,168,672,
        ]
dt      = 5
lambd   = [
        1,10,10,20,10
        ]
timestamps  = [        
        2,4,6,14,56
        ]

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(
        filepath,
        read_only   = True,
        keep_links  = False
        )
    return wb.sheetnames

sheets  = get_sheetnames_xlsx(filename)
sheets.remove(sheets[0])

for i in models:
    train   = sheets[i*3]
    test    = sheets[1 + i*3]
    sim     = sheets[2 + i*3]
    
# read Excel file into DataFrame
    df_train = pd.read_excel(
        filename,
        sheet_name  = train,
#        header      = 0,
        )
    
    df_test = pd.read_excel(
        filename,
        sheet_name  = test,
#        header      = 0,
        )
    
    df_sim = pd.read_excel(
        filename,
        sheet_name  = sim,
#        header      = 0,
        )

    for j in range(0,len(feature)):    
    
        t       = np.arange(0,len(df_test)*dt,dt)/60
        obs     = df_test[feature[j]]
        sim     = df_sim[feature[j]]
        
        fig     = plt.figure(
            dpi         = 100,
            )
        fig.set_figheight(12)
        fig.set_figwidth(18)
        
        title   = 'SINDY metamodel performance per element - ' + feature_type + ': ' + feature[j].capitalize()
        title   = title + '\nSP = ' + str(sp[i]) + '; dt = 5 min; Nu = 2%; FW = SP'
        
        fig.suptitle(
            title,
            fontsize = 20,
            )
        
        ax0 = plt.subplot2grid(
            shape       = (2, 3),
            loc         = (0, 0),
            colspan     = 2,
            )
        ax1 = plt.subplot2grid(
            shape       = (2, 3),
            loc         = (1, 0),
            colspan     = 2,
            )
        ax2 = plt.subplot2grid(
            shape       = (2, 3),
            loc         = (0, 2),
            colspan     = 1,
            )
        ax3 = plt.subplot2grid(
            shape       = (2, 3),
            loc         = (1, 2),
            colspan     = 1,
            )
    
        ax0.plot(
            t, obs,
            label = 'observed',
            )
        ax0.plot(
            t, sim,
            label = 'simulated'
            )
        ax0.set_title(
            'Comparison between observed and simulated values per timestamp',
            fontsize = 16,
            )
        ax0.legend(
            loc             = 'best',
        #    bbox_to_anchor  = (0.5, -0.35, 0, 0),
            ncol            = 2,
            prop={'size': 14},
            )
        ax0.set_xlabel(
            xlabel      = 'timestamps (hr)',
            fontsize    = 14,
            )
        ax0.set_ylabel(
            ylabel      = feature_type + ' ' + units,
            fontsize    = 14,
            )
        loc = ticker.MultipleLocator(base = timestamps[i])
        ax0.xaxis.set_major_locator(loc)
        
        ax1.plot(t,
                 sim - obs,
                 label = '(sim - obs)'
                 )
        ax1.set_title(
            'Feature error (simulate - observed) per timestamp',
            fontsize = 16,
            )
        ax1.legend(
            loc             = 'best',
        #    bbox_to_anchor  = (0.5, -0.35, 0, 0),
            ncol            = 2,
            prop={'size': 14},
            )
        ax1.set_xlabel(
            xlabel      = 'timestamps (hr)',
            fontsize    = 14,
            )
        ax1.set_ylabel(
            ylabel      = feature_type + ' error ' + units,
            fontsize    = 14,
            )
        loc = ticker.MultipleLocator(base = timestamps[i])
        ax1.xaxis.set_major_locator(loc)
        
        ax2.scatter(obs, sim)
        ax2.plot(
            obs, obs,
            ls      = '-',
            lw      = 2,
            color   = 'red',
            zorder  = 1,
            )
        ax2.set_title(
            'Scatter with simulated and observed values',
            fontsize = 16,
            )
        ax2.set_xlabel(
            xlabel      = feature_type + ' observed ' + units,
            fontsize    = 14,
            )
        ax2.set_ylabel(
            ylabel      = feature_type + ' simulated ' + units,
            fontsize    = 14,
            )
        
        ax3.hist(
            sim - obs,
            orientation     = 'horizontal',
            histtype        = 'bar',
            rwidth          = 2,
            )
        ax3.set_title(
            'Histogram of the feature error',
            fontsize = 16,
            )
        ax3.set_ylabel(
            ylabel      = feature_type + ' error\n (simulate - observed) ' + units,
            fontsize    = 14,
            )
        ax3.set_xlabel(
            xlabel      = 'Frequency (No.',
            fontsize    = 14,
            )
        ax3.text(
            0.95, 0.09,
            'Total predictions:\n' + str(round(sp[i]*60/dt,)) + ' und', 
            fontsize    = 14,
            ha          = 'right',
            va          = 'center',
            rotation    = 0,
            transform   = ax3.transAxes,
            )
        
        for ax in fig.axes:
            ax.tick_params(
            axis    = 'both',
            which   ='both',
            length  = 0,
            )
            ax.grid(
                True,
                axis    = 'both',
                which   ='both',
                ls      = "-",
                color   = '0.65'
                )
            ax.tick_params(
             axis        = 'both',
             which       = 'both',
             labelsize   = 14,
             )    
        
        plt.tight_layout(
            pad = 3.0,
            )
    
        plt.savefig('Plots/plots_' + feature[j].capitalize() + '_sp=' + str(sp[i]) + '_dt=5_λ=' + str(lambd[i]) + '_nu=2_sc_ls-1_fw=sp.png')
        plt.close() 
