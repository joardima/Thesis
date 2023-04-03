# -*- coding: utf-8 -*-

#%% GENERATE EQUATIONS IN EXCEL FILE

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import copy
import re
from openpyxl import load_workbook
plt.rcParams["font.family"] = "Times New Roman"

filename        = r'xlsx_files/twoloops_results_MATLAB_HDQ_equations.xlsx'

# read Excel file into DataFrame

features = pd.read_excel(
    filename,
    sheet_name  = 'features',
    header      = None,
    ).to_numpy()
features = features[:,1:].flatten()
features = np.reshape(
    features, (len(features), 1)
    )

theta = pd.read_excel(
    filename,
    sheet_name  = 'features',
    header      = None,
    ).to_numpy()

delete = '_' + '{' + '}' + "'" + "'"
string = '\u209c' + '\u208a' + '\u2081'
features = np.array([item + string for item in features])

sp = [24,48,72,168,672]
equations_df_all =[ ]
for s in sp:

    sn = 'Xi_' + str(s)

    Xi = pd.read_excel(
        filename,
        sheet_name  = sn,
        header      = None,
        ).to_numpy()
    
    Xi = Xi.T
        
    for i in delete:
        for j in range(0,len(features)):
            if i == '{':
                features[j,0] = re.sub(i, '(', features[j,0])
            if i == '}':
                features[j,0] = re.sub(i, ')', features[j,0])
            if i != '}' and i != '}':      
                features[j,0] = re.sub(i, '', features[j,0])
    for i in delete:
        for j in range(0,np.shape(theta)[1]):
            if i == '{':
                theta[0,j] = re.sub(i, '(', str(theta[0,j]))
            if i == '}':
                theta[0,j] = re.sub(i, ')', str(theta[0,j]))
            if i != '}' and i != '}':      
                theta[0,j] = re.sub(i, '', str(theta[0,j]))    
        
    equations       = np.array([ ])
    for i in range(0,len(Xi)):       
        equ         = np.array([ ])
        for j in range(0,np.shape(Xi)[1]):
            if j == 0:
                if Xi[i,j] != 0:                
                    temp    = str(round(Xi[i,j],2))
                else:
                    temp    = ''    
            else:
                if Xi[i,j] != 0:
                    temp    = str(round(Xi[i,j],2)) + 'x' + str(theta[0,j]) + '\u209c'
                else:
                    temp    = ''
            equ     = np.append(equ,temp)
        equations   = np.append(equations,equ)    
        
    equations = np.reshape(
        equations,
        (i + 1,
        len(equ))
        )
    
    equations = np.hstack(
        (features,equations)
        )
    
    equations = pd.DataFrame (equations)
 
    book = load_workbook(filename)
#    writer = pd.ExcelWriter(
        filename,
        engine = 'openpyxl'
        )
    writer.book = book
    
    equations.to_excel(
        writer,
        sheet_name = sn + '_equ',
        header  = False,
        index   = False,
        )
    writer.close()
    equations_df_all.append(equations)

#%% CREATE LIST FOR EXPORT EQUATIONS TO EXCEL

equ_list = [ ]
equations = equations_df_all[4]
for i in range(0,len(equations)):
    row = equations.iloc[i,:]
    temp = [ ]
    for j in range(0,len(row)):
        if j < 1:
            temp.append(row[j] + ' = ')
        if len(row[j]) > 0 and j == 1:
            temp.append(row[j])
        if len(row[j]) > 0 and j > 1:
            if row[j][0] == '-':
                temp.append(row[j])
            if row[j][0] != '-' and row[j][0] != '':
                temp.append('+' + row[j])
    temp = ''.join(temp)
    if temp[-1] == ' ':        
        equ_list.append(temp + '0')
    else:
        equ_list.append(temp)
    

#%% METADATA

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import copy
import re
from openpyxl import load_workbook
plt.rcParams["font.family"] = "Times New Roman"

filename        = r'xlsx_files/twoloops_results_MATLAB_HDQ_equations.xlsx'

sp          = [24,48,72,168,672]
dt          = [5,5,5,5,5]
lambd       = [1,10,10,20,5]
noise       = [2,2,2,2,2]
elements    = [6,6,8]
indexes     = [
    None,
    elements[0],
    elements[0] + elements[1],
    elements[0] + elements[1] + elements[2],
    ] 

# read Excel file into DataFrame

features = pd.read_excel(
    filename,
    sheet_name  = 'features',
    header      = None,
    ).to_numpy()
features = features[:,1:].flatten()
features = np.reshape(
    features, (len(features), 1)
    )

theta = pd.read_excel(
    filename,
    sheet_name  = 'features',
    header      = None,
    ).to_numpy()

delete = '_' + '{' + '}' + "'" + "'"
string = '\u209c' + '\u208a' + '\u2081'
features = np.array([item + string for item in features])

nz_coeff_number = [ ]; nz_coeff_value_min = [ ]; nz_coeff_value_max = [ ]; nz_coeff_len_min = [ ]; nz_coeff_len_max = [ ]
nz_coeff_number_pdq = [ ]; nz_coeff_value_min_pdq = [ ]; nz_coeff_value_max_pdq = [ ]; nz_coeff_len_min_pdq = [ ]; nz_coeff_len_max_pdq = [ ];
#coeff_pervariable_number = [ ]; coeff_pervariable_value_min = [ ]; coeff_pervariable_value_max = [ ];

for s in sp:
#    cpvn        = [ ];     cpqdmin     = [ ];   cpqdmax     = [ ]
    sn = 'Xi_' + str(s)

    Xi = pd.read_excel(
        filename,
        sheet_name  = sn,
        header      = None,
        )
    
    Xi = Xi.abs()    
    nzcn = Xi.astype(bool).sum(axis=0)
    nzcn = nzcn.sum()
    nz_coeff_number.append(nzcn)
    
    temp        = copy.deepcopy(Xi)
    temp.replace(0, np.nan, inplace=True)
    nz_coeff_value_min.append(temp.stack().min())
    nz_coeff_value_max.append(temp.stack().max())
    
    temp        = Xi.astype(bool).sum(axis=0)
    temp.replace(0, np.nan, inplace=True)
    nz_coeff_len_min.append(int(temp.min()))
    nz_coeff_len_max.append(int(temp.max()))
    
    nzcn_pdq = [ ]; nzcvmin_pdq = [ ]; nzcvmax_pdq = [ ]; nzclmin_pdq = [ ]; nzclmax_pdq = [ ];
    for i in range(0,len(indexes) - 1):
        
        feature        = Xi.iloc[
            :,
            indexes[i]:indexes[i + 1]
            ]
          
        nzcn = feature.astype(bool).sum(axis=0)
        nzcn = nzcn.sum()
        nzcn_pdq.append(nzcn)
        
        temp        = copy.deepcopy(feature)
        temp.replace(0, np.nan, inplace=True)
        nzcvmin_pdq.append(temp.stack().min())
        nzcvmax_pdq.append(temp.stack().max())
        
        temp        = feature.astype(bool).sum(axis=0)
        temp.replace(0, np.nan, inplace=True)
        nzclmin_pdq.append(temp.min())
        nzclmax_pdq.append(temp.max())
        
    nz_coeff_number_pdq.append(nzcn_pdq)
    nz_coeff_value_min_pdq.append(nzcvmin_pdq)
    nz_coeff_value_max_pdq.append(nzcvmax_pdq)
    nz_coeff_len_min_pdq.append(nzclmin_pdq)
    nz_coeff_len_max_pdq.append(nzclmax_pdq)
    
metadata                        = pd.DataFrame(sp, columns=['sp'])
metadata['dt']                  = dt
metadata['lambd']               = lambd
metadata['noise']               = noise

metadata['nz_coeff_number']     = nz_coeff_number
metadata['nz_coeff_value_min']  = nz_coeff_value_min
metadata['nz_coeff_value_max']  = nz_coeff_value_max
metadata['nz_coeff_len_min']    = nz_coeff_len_min
metadata['nz_coeff_len_max']    = nz_coeff_len_max

nz_coeff_number_pdq     = np.array(nz_coeff_number_pdq)
nz_coeff_value_min_pdq  = np.array(nz_coeff_value_min_pdq)
nz_coeff_value_max_pdq  = np.array(nz_coeff_value_max_pdq)
nz_coeff_len_min_pdq    = np.array(nz_coeff_len_min_pdq)
nz_coeff_len_max_pdq    = np.array(nz_coeff_len_max_pdq)
   
metadata['nz_coeff_number_p']       = nz_coeff_number_pdq[:,0]
metadata['nz_coeff_value_min_p']    = nz_coeff_value_min_pdq[:,0]   
metadata['nz_coeff_value_max_p']    = nz_coeff_value_max_pdq[:,0]
metadata['nz_coeff_len_min_p']      = nz_coeff_len_min_pdq[:,0]
metadata['nz_coeff_len_max_p']      = nz_coeff_len_max_pdq[:,0]   
        
metadata['nz_coeff_number_d']       = nz_coeff_number_pdq[:,1]
metadata['nz_coeff_value_min_d']    = nz_coeff_value_min_pdq[:,1]   
metadata['nz_coeff_value_max_d']    = nz_coeff_value_max_pdq[:,1]    
metadata['nz_coeff_len_min_d']      = nz_coeff_len_min_pdq[:,1]   
metadata['nz_coeff_len_max_d']      = nz_coeff_len_max_pdq[:,1]
       
metadata['nz_coeff_number_q']       = nz_coeff_number_pdq[:,2]
metadata['nz_coeff_value_min_q']    = nz_coeff_value_min_pdq[:,2]
metadata['nz_coeff_value_max_q']    = nz_coeff_value_max_pdq[:,2]
metadata['nz_coeff_len_min_q']      = nz_coeff_len_min_pdq[:,2]
metadata['nz_coeff_len_max_q']      = nz_coeff_len_max_pdq[:,2]

#%% BAR PLOT METADATA COEFFICIENTS (4 SUBPLOTS)

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import copy
import re
from openpyxl import load_workbook
plt.rcParams["font.family"] = "Times New Roman"

width = 0.25

fig, axs = plt.subplots(
        4,
        figsize     = (19, 10),
        dpi         = 100,
        sharex      = 'col',
        sharey      = 'row',
        )

title   = 'SINDY metamodel number of coefficients per SP\ndt = 5 min and Nu = 2%'

fig.suptitle(
    title,
    fontsize = 20,
    )
fig.supxlabel(
    'SINDY metamodel hyperparameters',
    fontsize    = 16,
    x = 0.5, y = 0.02,
    )
fig.supylabel(
    'Coefficients (number)',
    fontsize    = 16,
    x = 0.075, y = 0.5,
    )

col = [
    metadata.columns.get_loc('nz_coeff_number_p'),
    metadata.columns.get_loc('nz_coeff_number_d'),
    metadata.columns.get_loc('nz_coeff_number_q'),
    metadata.columns.get_loc('nz_coeff_number'),
    ]

subtitle = [
    'Presure head features',
    'Demand features',
    'Flow features',
    'SINDY metamodel',
    ]

for i in range(0,len(col)):

    x = np.arange(0,len(metadata['sp']))
    axs[i].bar(
        x,
        metadata.iloc[:,col[i]],
        width,
        label = '',
        zorder=2
        )

for i in range(0,len(metadata)):
    labels = 'SP=' + metadata['sp'].astype(str) + ' hr ; '
    labels = labels + 'λ=' + metadata['lambd'].astype(str)   

i = 0
for ax in axs.flatten():
    ax.tick_params(
        axis    = 'both',
        which   ='both',
        length  = 0,
        )
    ax.grid(
        True,
        axis    = 'both',
        which   ='both',
        ls      = "-",
        color   = '0.65'
        )
    ax.set_xticks(
        x, labels,
        fontsize    = 16,
        rotation    = 0,
        )
    
    ax.tick_params(
         axis        = 'both',
         which       = 'major',
         labelsize   = 16,
         )    

    ax.yaxis.set_major_locator(
             ticker.MultipleLocator(100)
             )
    ax.set_title(
        subtitle[i],
            fontdict={
                'fontsize': 16,
                }
        )
    i += 1

for ax in axs[:3].flatten():
    ax.set_ylim([
        0,
        200
        ])
    ax.yaxis.set_major_locator(
             ticker.MultipleLocator(50)
             )

plt.savefig('Plots/bar_coeff_sp_dt_λ_nu_all_sc_ls-1.png')
plt.close() 
    

#%% LINE PLOT METADATA COEFFICIENTS (4 SUBPLOTS)

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import copy
import re
from openpyxl import load_workbook
plt.rcParams["font.family"] = "Times New Roman"

width = 0.25

fig, axs = plt.subplots(
        4,
        figsize     = (19, 10),
        dpi         = 100,
        sharex      = 'col',
        sharey      = 'row',
        )

title   = 'SINDY metamodel number of coefficients per SP\ndt = 5 min and Nu = 2%'

fig.suptitle(
    title,
    fontsize = 20,
    )
fig.supxlabel(
    'SINDY metamodel hyperparameters',
    fontsize    = 16,
    x = 0.5, y = 0.02,
    )
fig.supylabel(
    'Coefficients (number)',
    fontsize    = 16,
    x = 0.075, y = 0.5,
    )

col = [
    metadata.columns.get_loc('nz_coeff_number_p'),
    metadata.columns.get_loc('nz_coeff_number_d'),
    metadata.columns.get_loc('nz_coeff_number_q'),
    metadata.columns.get_loc('nz_coeff_number'),
    ]

subtitle = [
    'Presure head features',
    'Demand features',
    'Flow features',
    'SINDY metamodel',
    ]

for i in range(0,len(col)):

    x = np.arange(0,len(metadata['sp']))
    axs[i].plot(
        x,
        metadata.iloc[:,col[i]],
#        width,
        label = '',
        zorder=2
        )

for i in range(0,len(metadata)):
    labels = 'SP=' + metadata['sp'].astype(str) + ' hr ; '
    labels = labels + 'λ=' + metadata['lambd'].astype(str)   

i = 0
for ax in axs.flatten():
    ax.tick_params(
        axis    = 'both',
        which   ='both',
        length  = 0,
        )
    ax.grid(
        True,
        axis    = 'both',
        which   ='both',
        ls      = "-",
        color   = '0.65'
        )
    ax.set_xticks(
        x, labels,
        fontsize    = 16,
        rotation    = 0,
        )
    
    ax.tick_params(
         axis        = 'both',
         which       = 'major',
         labelsize   = 16,
         )    

    ax.yaxis.set_major_locator(
             ticker.MultipleLocator(100)
             )
    ax.set_title(
        subtitle[i],
            fontdict={
                'fontsize': 16,
                }
        )
    i += 1

for ax in axs[:3].flatten():
    ax.set_ylim([
        0,
        200
        ])
    ax.yaxis.set_major_locator(
             ticker.MultipleLocator(50)
             )

plt.savefig('Plots/line_coeff_sp_dt_λ_nu_all_sc_ls-1.png')
plt.close() 
    
#%% LINE PLOT METADATA COEFFICIENTS (1 PLOT)

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import copy
import re
from openpyxl import load_workbook
plt.rcParams["font.family"] = "Times New Roman"

width = 0.25

fig, axs = plt.subplots(
        1,
        figsize     = (19, 10),
        dpi         = 100,
        sharex      = 'col',
        sharey      = 'row',
        )

title   = 'SINDY metamodel number of coefficients per SP\ndt = 5 min and Nu = 2%'

fig.suptitle(
    title,
    fontsize = 20,
    )
fig.supxlabel(
    'SINDY metamodel hyperparameters',
    fontsize    = 16,
    x = 0.5, y = 0.02,
    )
fig.supylabel(
    'Coefficients (number)',
    fontsize    = 16,
    x = 0.075, y = 0.5,
    )

col = [
    metadata.columns.get_loc('nz_coeff_number_p'),
    metadata.columns.get_loc('nz_coeff_number_d'),
    metadata.columns.get_loc('nz_coeff_number_q'),
    metadata.columns.get_loc('nz_coeff_number'),
    ]

subtitle = [
    'Presure head',
    'Demand',
    'Flow',
    'SINDY metamodel',
    ]

for i in range(0,len(col)):

    x = np.arange(0,len(metadata['sp']))
    axs.plot(
        x,
        metadata.iloc[:,col[i]],
        label = subtitle[i],
        zorder      = 2,
        linewidth   = 2,
        markersize  = 10,
        marker  = 'o',
        )

for i in range(0,len(metadata)):
    labels = 'SP=' + metadata['sp'].astype(str) + ' hr ; '
    labels = labels + 'λ=' + metadata['lambd'].astype(str)   

i = 0

axs.tick_params(
    axis    = 'both',
    which   ='both',
    length  = 0,
    )
axs.grid(
    True,
    axis    = 'y',
    which   ='both',
    ls      = "-",
    color   = '0.65'
    )
axs.set_xticks(
    x, labels,
    fontsize    = 16,
    rotation    = 0,
    )

axs.tick_params(
     axis        = 'both',
     which       = 'major',
     labelsize   = 16,
     )    

axs.yaxis.set_major_locator(
         ticker.MultipleLocator(50)
         )

axs.legend(
    loc             = 'best',
#    bbox_to_anchor  = (0.5, -0.35, 0, 0),
    ncol            = 4,
    prop={'size': 14},
    )

    
plt.savefig('Plots/line_all_coeff_sp_dt_λ_nu_all_sc_ls-1.png')
plt.close() 
    



    
 

    



    
 

